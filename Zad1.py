import openpyxl

# Create a new workbook
workbook = openpyxl.Workbook()

# Create the first sheet for the first file

sheet2 = workbook.create_sheet(title="1.txt")
sheet2 = workbook.create_sheet(title="2.txt")
sheet3 = workbook.create_sheet(title="3.txt")
sheet4 = workbook.create_sheet(title="4.txt")
sheet5 = workbook.create_sheet(title="5.txt")
sheet2 = workbook.create_sheet(title="6.txt")
sheet2 = workbook.create_sheet(title="7.txt")
sheet3 = workbook.create_sheet(title="8.txt")
sheet4 = workbook.create_sheet(title="9.txt")
sheet5 = workbook.create_sheet(title="10.txt")



# Save the workbook
workbook.save("bajzel.xlsx")
