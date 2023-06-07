import openpyxl
from openpyxl.styles import Font


workbook = openpyxl.Workbook()
sheet = workbook.active
table_size = 10

#naglowki kolumnowe
for i in range(1, 11):
    sheet.cell(row=1, column=i+1, value=i)
    cell = sheet.cell(row=1, column=i+1)
    font = Font(bold=True)
    cell.font=font

#naglowki wierszowe
for i in range(1, 11):
    sheet.cell(row=i+1, column=1, value=i)
    cell = sheet.cell(row=i+1, column=1)
    font = Font(bold=True)
    cell.font=font

#policzenie i wpisanie tabliczki mnozenia
for i in range(1, table_size + 1):
    for j in range(1, table_size + 1):
        result = i * j
        sheet.cell(row=i+1, column=j+1, value=result)

        

workbook.save("tabliczka.xlsx")


